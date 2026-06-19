"""Validation report — kategori/severity dağılımı + sistemik vs tekil tespiti."""
from __future__ import annotations

from collections import Counter
from typing import Any

from app.core.config import settings
from app.features.validation.models import (
    Issue,
    ValidationResult,
)


def category_breakdown(results: dict[int, ValidationResult]) -> dict[str, int]:
    c: Counter[str] = Counter()
    for r in results.values():
        for i in r.issues:
            c[i.category.value] += 1
    return dict(c)


def severity_breakdown(results: dict[int, ValidationResult]) -> dict[str, int]:
    s: Counter[str] = Counter()
    for r in results.values():
        for i in r.issues:
            s[i.severity.value] += 1
    return dict(s)


def rule_breakdown(results: dict[int, ValidationResult]) -> dict[str, int]:
    r: Counter[str] = Counter()
    for res in results.values():
        for i in res.issues:
            r[i.rule_id] += 1
    return dict(r)


def systemic_vs_unique(
    results: dict[int, ValidationResult],
    threshold_ratio: float | None = None,
) -> dict[str, list[str]]:
    if threshold_ratio is None:
        threshold_ratio = settings.validation_systemic_ratio
    total = max(len(results), 1)
    rb = rule_breakdown(results)
    systemic: list[str] = []
    unique: list[str] = []
    for rule_id, count in rb.items():
        if (count / total) >= threshold_ratio:
            systemic.append(rule_id)
        else:
            unique.append(rule_id)
    return {"systemic": sorted(systemic), "unique": sorted(unique)}


def issue_to_dict(issue: Issue) -> dict[str, Any]:
    return {
        "rule_id": issue.rule_id,
        "category": issue.category.value,
        "severity": issue.severity.value,
        "fields": list(issue.fields),
        "message": issue.message,
        "suggested_action": issue.suggested_action.value,
    }


def record_payload(record_id: int, result: ValidationResult) -> dict[str, Any]:
    return {
        "record_id": record_id,
        "issues": [issue_to_dict(i) for i in result.issues],
        "record_status": result.status.value,
    }


def report_xlsx(results: dict[int, ValidationResult]) -> bytes:
    """Validation sonuçlarını `.xlsx` olarak serileştir (3 sayfa: Özet, Issues, Sistemik-Tekil).

    `openpyxl` lazy import edilir (yalnız export çağrısında yüklenir).
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    bold = Font(bold=True)

    wb = Workbook()
    # 1) Özet
    ws = wb.active
    ws.title = "Özet"
    ws.append(["Metrik", "Değer"])
    for cell in ws[1]:
        cell.font = bold
    status_c: Counter[str] = Counter(r.status.value for r in results.values())
    ws.append(["Toplam kayıt", len(results)])
    for k in ("valid", "suspect", "rejected"):
        ws.append([f"status: {k}", status_c.get(k, 0)])
    for k, v in sorted(category_breakdown(results).items()):
        ws.append([f"category: {k}", v])
    for k, v in sorted(severity_breakdown(results).items()):
        ws.append([f"severity: {k}", v])
    for k, v in sorted(rule_breakdown(results).items()):
        ws.append([f"rule: {k}", v])

    # 2) Issues (kayıt başına her issue bir satır)
    ws2 = wb.create_sheet("Issues")
    ws2.append(
        [
            "record_id", "rule_id", "category", "severity",
            "fields", "message", "suggested_action", "record_status",
        ]
    )
    for cell in ws2[1]:
        cell.font = bold
    for rid, res in results.items():
        for issue in res.issues:
            ws2.append(
                [
                    rid,
                    issue.rule_id,
                    issue.category.value,
                    issue.severity.value,
                    ", ".join(issue.fields),
                    issue.message,
                    issue.suggested_action.value,
                    res.status.value,
                ]
            )

    # 3) Sistemik vs tekil
    ws3 = wb.create_sheet("Sistemik-Tekil")
    ws3.append(["tip", "rule_id"])
    for cell in ws3[1]:
        cell.font = bold
    sv = systemic_vs_unique(results)
    for rid in sv["systemic"]:
        ws3.append(["sistemik", rid])
    for rid in sv["unique"]:
        ws3.append(["tekil", rid])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def full_report(results: dict[int, ValidationResult]) -> dict[str, Any]:
    return {
        "summary": {
            "total_records": len(results),
            "by_category": category_breakdown(results),
            "by_severity": severity_breakdown(results),
            "by_rule": rule_breakdown(results),
        },
        "systemic_vs_unique": systemic_vs_unique(results),
        "records": [record_payload(rid, r) for rid, r in results.items()],
    }
